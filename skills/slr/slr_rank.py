"""SLR embedding + ranking step (run AFTER slr_scopus.py).

Loads the XLSX from slr_scopus.py, computes embeddings for all papers,
ranks by cosine similarity against the input text, and updates the XLSX
with scores and reordered ranks.

Usage:
  PYTHONPATH=src .venv/bin/python skills/slr/slr_rank.py \\
      --xlsx output/<stem>/SLR_Scopus_<tag>/SLR_Scopus_<tag>.xlsx \\
      --input <stem> [--top-n <N>] [--threshold <float>]
"""
import argparse
from pathlib import Path

import dotenv
dotenv.load_dotenv()

import yaml
import openpyxl

from research_finder.embedder import get_embedder
from research_finder.ranker import rank_papers
from research_finder.input_store import InputStore
from research_finder.models import Paper

# ── Args ───────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="SLR embed + rank step")
parser.add_argument("--xlsx", required=True, help="Path to the SLR xlsx from slr_scopus.py")
parser.add_argument("--input", dest="input_name", required=True,
                    help="Cached input name (stem or filename in input/embeddings/)")
parser.add_argument("--top-n", dest="top_n", type=int, default=None,
                    help="Top-N papers to keep after ranking (default: all)")
parser.add_argument("--threshold", dest="threshold", type=float, default=0.0,
                    help="Minimum cosine similarity score (default 0.0)")
args = parser.parse_args()

# ── Config ─────────────────────────────────────────────────────────────────
with open("config.yaml") as f:
    cfg = yaml.safe_load(f)
cfg["embedding"]["provider"] = "api"

# ── Load input embedding ───────────────────────────────────────────────────
store = InputStore(cfg)
text_emb, TEXT = store.load(args.input_name)
print(f"[input] '{args.input_name}': {len(TEXT)} chars, emb={text_emb.shape}")

# ── Load papers from XLSX ──────────────────────────────────────────────────
xlsx_path = Path(args.xlsx)
if not xlsx_path.exists():
    print(f"XLSX not found: {xlsx_path}")
    import sys
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active
headers = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

required = {"title", "abstract", "authors", "year", "journal", "doi",
            "scopus_id", "citations", "retrieval_query", "topic_label",
            "publication", "issn", "volume", "issue", "pages", "doc_type",
            "open_access", "author_keywords", "affiliations", "source", "url"}
missing = required - set(idx)
if missing:
    print(f"XLSX missing columns: {missing}")

papers: list[Paper] = []
rows_data: list[list] = []

for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c+1).value for c in range(len(headers))]
    p = Paper(
        title=row_vals[idx["title"]] or "",
        abstract=row_vals[idx["abstract"]] or None,
        authors=(row_vals[idx["authors"]] or "").split("; ") if row_vals[idx["authors"]] else [],
        year=row_vals[idx["year"]] or None,
        publication=row_vals[idx["journal"]] or None,
        doi=row_vals[idx["doi"]] or None,
        scopus_id=row_vals[idx["scopus_id"]] or None,
        retrieval_query=row_vals[idx["retrieval_query"]] or "",
    )
    papers.append(p)
    rows_data.append(row_vals)

print(f"Loaded {len(papers)} papers from XLSX")

# ── Embed + rank ───────────────────────────────────────────────────────────
print("=" * 80)
print("Embedding & ranking")
print("=" * 80)

embedder = get_embedder(cfg)
print(f"Text embedding: {text_emb.shape}")

paper_texts = [f"{p.title} {p.abstract or ''}" for p in papers]
paper_embs = embedder.embed_batch(paper_texts)
print(f"Paper embeddings: {paper_embs.shape}")

top_n = args.top_n if args.top_n else len(papers)
ranked = rank_papers(text_emb, paper_embs, papers, top_n=top_n, threshold=args.threshold)
print(f"Ranked {len(ranked)} papers (threshold={args.threshold})")

# Print top 10 preview
for i, p in enumerate(ranked[:10], 1):
    print(f"  #{i:>3}  [{p.score:.4f}]  {p.title[:80]}")
if len(ranked) > 10:
    print(f"  ... and {len(ranked) - 10} more")
print()

# ── Update XLSX with ranks and scores ──────────────────────────────────────
print(f"Updating XLSX: {xlsx_path}")

# Build lookup from paper identity → row index
# Use scopus_id or doi as key
key_to_row: dict[str, int] = {}
for ridx, row in enumerate(rows_data):
    key = row[idx["scopus_id"]] or row[idx["doi"]] or row[idx["title"]]
    if key:
        key_to_row[str(key).strip()] = ridx

rank_col = idx["rank"] + 1
score_col = idx["score"] + 1

# Reset all ranks/scores first
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=rank_col, value=0)
    ws.cell(row=r, column=score_col, value=0.0)

for new_rank, p in enumerate(ranked, 1):
    key = p.scopus_id or p.doi or p.title
    key = str(key).strip()
    if key in key_to_row:
        row_num = key_to_row[key] + 2  # +2: 0-based → 1-based + header
        ws.cell(row=row_num, column=rank_col, value=new_rank)
        ws.cell(row=row_num, column=score_col, value=round(p.score, 6))

wb.save(xlsx_path)
print(f"Updated {len(ranked)} papers with scores and ranks.")
print("Done.")
print()
print("Next: post-SLR analysis — keyword verification → journal exclusion → 10 analyses")
print("  see skills/test-slr-analysis/SKILL.md")
