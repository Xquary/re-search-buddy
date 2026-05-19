"""Recover missing abstracts in an SLR xlsx file via Scopus Abstract Retrieval API.

Reads scopus_ids from the xlsx, fetches abstracts for empty rows, and saves
the updated file. Idempotent — skips papers that already have abstracts.

Usage:
  PYTHONPATH=src .venv/bin/python skills/slr/recover_abstracts.py \
      "output/Seeds of Green_Methodology/SLR_Scopus_q1_max2000_y2016-2021/SLR_Scopus_q1_max2000_y2016-2021.xlsx"
"""
import argparse
import asyncio
import os
import sys
from pathlib import Path

import dotenv
import httpx
from openpyxl import load_workbook

dotenv.load_dotenv()

_ABSTRACT_URL = "https://api.elsevier.com/content/abstract/scopus_id/{}"


def main():
    parser = argparse.ArgumentParser(description="Recover missing abstracts from SLR xlsx")
    parser.add_argument("xlsx", help="Path to the SLR xlsx file")
    parser.add_argument("--delay", type=float, default=0.15,
                        help="Seconds between API requests (default 0.15)")
    parser.add_argument("--progress-every", type=int, default=100,
                        help="Print progress every N papers (default 100)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}")
        sys.exit(1)

    # Read xlsx
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Locate columns
    try:
        sid_col = headers.index("scopus_id")
        abs_col = headers.index("abstract")
    except ValueError as e:
        print(f"Error: required column missing — {e}")
        sys.exit(1)
    lang_col = headers.index("language") if "language" in headers else None

    # Find rows that need abstracts or language
    to_fetch: list[tuple[int, str]] = []  # (row_idx, scopus_id)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        sid = row[sid_col]
        abstract = row[abs_col]
        language = row[lang_col] if lang_col is not None else None
        need_abstract = sid and str(sid).strip() and not (abstract and str(abstract).strip())
        need_language = lang_col is not None and sid and str(sid).strip() and not (language and str(language).strip())
        if need_abstract or need_language:
            to_fetch.append((row_idx, str(sid).strip()))

    if not to_fetch:
        print("All papers already have abstracts — nothing to do.")
        return

    total = len(to_fetch)
    print(f"Found {total} papers without abstracts")

    api_key = os.environ.get("SCOPUS_API_KEY", "")
    if not api_key:
        print("Error: SCOPUS_API_KEY not set in environment")
        sys.exit(1)

    headers_http = {
        "X-ELS-APIKey": api_key,
        "Accept": "application/json",
    }

    fetched = 0
    failed = 0
    empty = 0
    errors: list[tuple[str, str]] = []

    async def run():
        nonlocal fetched, failed, empty, errors
        async with httpx.AsyncClient(
            headers=headers_http, timeout=30.0, follow_redirects=True
        ) as client:
            for i, (row_idx, sid) in enumerate(to_fetch):
                try:
                    resp = await client.get(
                        _ABSTRACT_URL.format(sid),
                        params={"view": "FULL"},
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    root = (
                        data.get("abstracts-retrieval-response")
                        or data.get("abstract-retrieval-response")
                        or {}
                    )
                    coredata = root.get("coredata", {})
                    abstract = coredata.get("dc:description")
                    if abstract:
                        ws.cell(row=row_idx, column=abs_col + 1, value=abstract)
                        fetched += 1
                    else:
                        empty += 1
                    # Also extract language if column exists
                    if lang_col is not None:
                        lang = root.get("language", {}).get("@xml:lang")
                        if lang:
                            ws.cell(row=row_idx, column=lang_col + 1, value=lang)
                except Exception as e:
                    failed += 1
                    if len(errors) < 3:
                        errors.append((sid, str(e)[:120]))

                if (i + 1) % args.progress_every == 0:
                    print(f"  {i+1}/{total}  (ok={fetched} failed={failed} empty={empty})")

                await asyncio.sleep(args.delay)

        print(f"\nDone: {fetched} fetched, {empty} empty, {failed} failed (of {total})")
        if errors:
            print("First errors:")
            for sid, msg in errors:
                print(f"  {sid}: {msg}")
        if failed >= total * 0.9 and total > 10:
            print(f"WARNING: {failed}/{total} requests failed — "
                  f"check API key, quota, and institutional IP entitlements")

    asyncio.run(run())

    # Save
    wb.save(str(xlsx_path))
    print(f"Saved: {xlsx_path}")


if __name__ == "__main__":
    main()
