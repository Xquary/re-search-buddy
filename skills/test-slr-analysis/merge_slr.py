"""Merge 2016-2021 and 2022-2026 SLR xlsx files, deduplicate, re-rank, run keyword verification."""
import openpyxl
import numpy as np
from pathlib import Path
from collections import Counter, OrderedDict
import re

OLD_XLSX = "output/Seeds of Green_Methodology/SLR_Scopus_q1_max2000/SLR_Scopus_q1_max2000.xlsx"
NEW_XLSX = "output/Seeds of Green_Methodology/SLR_Scopus_q1_max2000_y2016-2021/SLR_Scopus_q1_max2000_y2016-2021.xlsx"
OUT_DIR = Path("output/Seeds of Green_Methodology/SLR_Scopus_ABM_Energy_2016-2026_qcomb_max2000")
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_XLSX = OUT_DIR / "SLR_Scopus_ABM_Energy_2016-2026_qcomb_max2000.xlsx"

# ── 1. Load and merge SLR sheets ──
print("Loading xlsx files...")
wb_old = openpyxl.load_workbook(OLD_XLSX)
ws_old = wb_old["SLR"]
wb_new = openpyxl.load_workbook(NEW_XLSX)
ws_new = wb_new["SLR"]

# Read headers
h_old = [c.value for c in ws_old[1]]
h_new = [c.value for c in ws_new[1]]
assert h_old == h_new, f"Headers don't match: {h_old} vs {h_new}"
headers = h_old

# Read all papers
def read_sheet(ws):
    papers = []
    for row_idx in range(2, ws.max_row + 1):
        row = {}
        for col_idx, h in enumerate(headers):
            row[h] = ws.cell(row=row_idx, column=col_idx + 1).value
        papers.append(row)
    return papers

papers_old = read_sheet(ws_old)
papers_new = read_sheet(ws_new)
print(f"Loaded: {len(papers_old)} (2022-2026) + {len(papers_new)} (2016-2021) = {len(papers_old) + len(papers_new)}")

# ── 2. Deduplicate by scopus_id, keeping highest score ──
seen = OrderedDict()  # scopus_id -> paper
dup_count = 0
for p in papers_old + papers_new:
    sid = str(p.get("scopus_id", "")).strip()
    if not sid:
        # Fallback: use DOI
        sid = str(p.get("doi", "")).strip()
    if not sid:
        # Last resort: use title
        sid = str(p.get("title", "")).strip().lower()[:100]
    if sid in seen:
        # Keep the one with higher score
        if (p.get("score") or 0) > (seen[sid].get("score") or 0):
            seen[sid] = p
        dup_count += 1
    else:
        seen[sid] = p

all_papers = list(seen.values())
# Sort by score descending
all_papers.sort(key=lambda p: p.get("score") or 0, reverse=True)
print(f"After dedup: {len(all_papers)} papers (removed {dup_count} duplicates)")

# ── 3. Re-rank ──
for i, p in enumerate(all_papers):
    p["rank"] = i + 1

# ── 4. Write merged SLR sheet ──
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "SLR"

# Write headers
for col_idx, h in enumerate(headers, 1):
    ws_out.cell(row=1, column=col_idx, value=h)

# Write papers
for row_idx, p in enumerate(all_papers, 2):
    for col_idx, h in enumerate(headers, 1):
        ws_out.cell(row=row_idx, column=col_idx, value=p.get(h))

print(f"SLR sheet written: {len(all_papers)} rows")

# ── 5. Keyword verification ──
KEYWORD_SETS = {
    "D-energy": [
        "energy transition", "energy market", "energy system",
        "energy policy", "energy behaviour", "energy behavior",
    ],
    "E-industrial": [
        "industrial decarbonization", "industrial transition",
        "industrial ecology", "industrial reorganization",
        "industrial reorganisation",
    ],
    "F-green": [
        "green transition", "sustainable transition",
        "sustainability transition", "low-carbon transition",
        "green transformation",
    ],
    "G-firm&decision": [
        "firm", "firms", "corporate", "firm-level",
        "firm decision", "firm decisions", "firm behaviour",
        "firm behavior", "firm-level decision",
    ],
}

# Build keyword_hit column
keyword_hit_col = []
paper_subsets = []
abstract_idx = headers.index("abstract")
title_idx = headers.index("title")

for p in all_papers:
    title = str(p.get("title") or "").lower()
    abstract = str(p.get("abstract") or "").lower()
    text = f"{title} {abstract}"

    hits = []
    for prefix, terms in KEYWORD_SETS.items():
        for term in terms:
            if term.lower() in text:
                hits.append(f"{prefix}:{term}")

    keyword_hit = "; ".join(hits) if hits else ""
    keyword_hit_col.append(keyword_hit)

    subsets = set()
    for h in hits:
        if h.startswith("D-energy:"): subsets.add("energy")
        elif h.startswith("E-industrial:"): subsets.add("industrial")
        elif h.startswith("F-green:"): subsets.add("green")
        elif h.startswith("G-firm&decision:"): subsets.add("firm")
    paper_subsets.append(subsets)

# ── 6. Create keyword_verified, missing_topic_only, filtered_out (not matched) ──
# For consistency with original: keyword_verified = papers with any keyword hit
kw_papers = [(i, p, kw, ss) for i, (p, kw, ss) in enumerate(zip(all_papers, keyword_hit_col, paper_subsets)) if kw]
mt_papers = [(i, p, kw, ss) for i, (p, kw, ss) in enumerate(zip(all_papers, keyword_hit_col, paper_subsets)) if not kw]

print(f"\nKeyword verification:")
print(f"  keyword_verified: {len(kw_papers)} papers")
print(f"  missing_topic_only: {len(mt_papers)} papers")

# Also: papers that were filtered out by keyword scan (none in this merge approach)
filtered_out = []  # not applicable here

def write_sheet(wb, name, papers_list, extra_cols=None):
    ws = wb.create_sheet(title=name)
    # Write all original headers + extra
    all_h = list(headers) + (extra_cols or [])
    for col_idx, h in enumerate(all_h, 1):
        ws.cell(row=1, column=col_idx, value=h)

    for row_idx, (orig_idx, p, *extra_vals) in enumerate(papers_list, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=p.get(h))
        if extra_cols:
            for col_idx, val in enumerate(extra_vals, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

write_sheet(wb_out, "keyword_verified", [(i, p, kw) for i, p, kw, ss in kw_papers], ["keyword_hit"])
write_sheet(wb_out, "missing_topic_only", [(i, p) for i, p, kw, ss in mt_papers])

# ── 7. Create clean sheets (remove Energies + Sustainability Switzerland) ──
BAD_JOURNALS = {"Energies", "Sustainability Switzerland"}
journal_idx = headers.index("journal")

def create_clean_sheet(wb_out, source_name):
    ws_src = wb_out[source_name]
    has_kw = "keyword_hit" in [c.value for c in ws_src[1]]
    src_headers = [c.value for c in ws_src[1]]
    target_name = source_name + "_clean"
    ws_tgt = wb_out.create_sheet(title=target_name)

    # Copy headers
    for col_idx, h in enumerate(src_headers, 1):
        ws_tgt.cell(row=1, column=col_idx, value=h)

    removed = 0
    tgt_row = 2
    for row_idx in range(2, ws_src.max_row + 1):
        journal = str(ws_src.cell(row=row_idx, column=journal_idx + 1).value or "").strip()
        if journal in BAD_JOURNALS:
            removed += 1
            continue
        for col_idx in range(1, len(src_headers) + 1):
            ws_tgt.cell(row=tgt_row, column=col_idx, value=ws_src.cell(row=row_idx, column=col_idx).value)
        tgt_row += 1
    print(f"  {target_name}: removed {removed}, kept {tgt_row - 2}")

create_clean_sheet(wb_out, "keyword_verified")
create_clean_sheet(wb_out, "missing_topic_only")

# ── 8. Save ──
wb_out.save(OUT_XLSX)
print(f"\nSaved: {OUT_XLSX}")
