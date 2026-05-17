"""Universal post-SLR analysis for a single-sheet ranked xlsx.

Organizes output into 3 phase subdirs with numbered filenames:
  1_profiling/   — yearly trend, top journals, topic-label distribution
  2_content/     — keyword network, topic modeling, geographic distribution
  3_visuals/     — wordcloud(s)

Each chart is prefixed with its phase.subphase number (e.g. 1.1_yearly_trend.png).

Usage: PYTHONPATH=src .venv/bin/python skills/test-slr-analysis/analysis_universal.py <path/to/xlsx>
"""
import sys, re
from pathlib import Path
from collections import Counter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import NMF
from sklearn.manifold import TSNE

XLSX = Path(sys.argv[1])
EXCLUDE_JOURNALS = set()
if len(sys.argv) > 2 and sys.argv[2] == "--exclude-journals":
    EXCLUDE_JOURNALS = {j.strip().lower() for j in sys.argv[3].split("|") if j.strip()}
BASE = XLSX.parent / "charts" / "analysis"
P1 = BASE / "1_profiling";  P1.mkdir(parents=True, exist_ok=True)
P2 = BASE / "2_content";    P2.mkdir(parents=True, exist_ok=True)
P3 = BASE / "3_visuals";    P3.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
headers = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

rows = []
for r in range(2, ws.max_row + 1):
    rows.append([ws.cell(row=r, column=i+1).value for i in range(len(headers))])

def col(row, name):
    return row[idx[name]] if name in idx else None

raw_n = len(rows)
if EXCLUDE_JOURNALS:
    ji = idx.get("journal")
    rows = [r for r in rows if str(r[ji] or "").strip().lower() not in EXCLUDE_JOURNALS]
    print(f"[filter] excluded journals: {sorted(EXCLUDE_JOURNALS)}  ({raw_n - len(rows)} papers removed)")
print(f"[loaded] {len(rows)} papers from {XLSX.name}")
has_topic = "topic_label" in idx
print(f"[topics] topic_label column present: {has_topic}")

# ============================================================
# Phase 1 — Data Profiling
# ============================================================

# 1.1 Yearly trend (stacked by topic_label if present)
years = []
year_topic = []
for r in rows:
    y = col(r, "year")
    try: yi = int(str(y)[:4])
    except: continue
    years.append(yi)
    year_topic.append((yi, str(col(r, "topic_label") or "unlabelled")))

yc = Counter(years)
xs = sorted(yc.keys())
fig, ax = plt.subplots(figsize=(9, 4))
if has_topic:
    topics = sorted({t for _, t in year_topic})
    cmap = plt.get_cmap("tab10")
    bottom = np.zeros(len(xs))
    for ti, t in enumerate(topics):
        vals = np.array([sum(1 for yy, tt in year_topic if yy == x and tt == t) for x in xs])
        ax.bar(xs, vals, bottom=bottom, label=t, color=cmap(ti))
        bottom += vals
    ax.legend(fontsize=8)
else:
    ax.bar(xs, [yc[x] for x in xs], color="#2196F3")
ax.set_xlabel("Year"); ax.set_ylabel("Papers"); ax.set_title(f"1.1 Publications per year (n={sum(yc.values())})")
plt.tight_layout(); plt.savefig(P1 / "1.1_yearly_trend.png", dpi=130); plt.close()
print(f"  1.1 → yearly_trend.png  ({len(xs)} years)")

# 1.2 Top journals
journals = Counter()
for r in rows:
    j = col(r, "journal") or col(r, "publication") or ""
    if j: journals[str(j).strip()] += 1
top = journals.most_common(15)
fig, ax = plt.subplots(figsize=(9, 6))
ax.barh([j for j, _ in top[::-1]], [c for _, c in top[::-1]], color="#FF9800")
ax.set_xlabel("Papers"); ax.set_title("1.2 Top 15 journals")
plt.tight_layout(); plt.savefig(P1 / "1.2_top_journals.png", dpi=130); plt.close()
print(f"  1.2 → top_journals.png  ({len(journals)} unique journals)")

# 1.3 Topic-label distribution (if available)
if has_topic:
    tc = Counter(str(col(r, "topic_label") or "unlabelled") for r in rows)
    items = tc.most_common()
    fig, ax = plt.subplots(figsize=(7, 4))
    ax.bar([t for t, _ in items], [n for _, n in items], color="#9C27B0")
    for i, (_, n) in enumerate(items):
        ax.text(i, n + 0.1, str(n), ha="center", fontsize=9)
    ax.set_ylabel("Papers"); ax.set_title("1.3 Papers per topic label")
    plt.tight_layout(); plt.savefig(P1 / "1.3_topic_distribution.png", dpi=130); plt.close()
    print(f"  1.3 → topic_distribution.png  ({len(tc)} topics)")

# ============================================================
# Phase 2 — Content Analysis
# ============================================================

docs = []
for r in rows:
    parts = [col(r, "title") or "", col(r, "abstract") or "", col(r, "author_keywords") or ""]
    docs.append(" ".join(str(p) for p in parts))

vec = TfidfVectorizer(stop_words="english", max_features=300, ngram_range=(1,2),
                      min_df=2, max_df=0.7)
X = vec.fit_transform(docs)
terms = vec.get_feature_names_out()
print(f"  [tfidf] {X.shape[0]} docs × {X.shape[1]} terms")

n_topics = 6
nmf = NMF(n_components=n_topics, init="nndsvd", random_state=42, max_iter=400)
W = nmf.fit_transform(X); H = nmf.components_
topic_terms = []
for ti, comp in enumerate(H):
    top_idx = comp.argsort()[::-1][:8]
    topic_terms.append([terms[i] for i in top_idx])

# 2.1 Keyword network (NMF topic-term table)
fig, ax = plt.subplots(figsize=(11, 4))
ax.axis("off")
tbl = ax.table(cellText=[[f"T{i+1}"] + tt for i, tt in enumerate(topic_terms)],
               colLabels=["Topic"] + [f"#{i+1}" for i in range(8)], loc="center")
tbl.auto_set_font_size(False); tbl.set_fontsize(9); tbl.scale(1, 1.5)
ax.set_title("2.1 NMF topics (top terms)")
plt.tight_layout(); plt.savefig(P2 / "2.1_keyword_network_topics.png", dpi=130); plt.close()
print(f"  2.1 → keyword_network_topics.png  ({n_topics} topics)")

# 2.2 Topic modeling: t-SNE + topic-year trend
dom_topic = W.argmax(axis=1)
try:
    perp = min(30, max(5, len(rows)//4))
    coords = TSNE(n_components=2, perplexity=perp, random_state=42, init="pca").fit_transform(W)
    fig, ax = plt.subplots(figsize=(9, 7))
    cmap = plt.get_cmap("tab10")
    for t in range(n_topics):
        m = dom_topic == t
        ax.scatter(coords[m,0], coords[m,1], color=cmap(t),
                   label=f"T{t+1}: {', '.join(topic_terms[t][:3])}", s=40, alpha=0.75)
    ax.legend(fontsize=8, loc="best"); ax.set_title("2.2 t-SNE of NMF topic loadings")
    ax.set_xticks([]); ax.set_yticks([])
    plt.tight_layout(); plt.savefig(P2 / "2.2_topic_tsne.png", dpi=130); plt.close()
    print(f"  2.2 → topic_tsne.png")
except Exception as e:
    print(f"  ! t-SNE failed: {e}")

topic_year = {t: Counter() for t in range(n_topics)}
for r, t in zip(rows, dom_topic):
    y = col(r, "year")
    try: yy = int(str(y)[:4])
    except: continue
    topic_year[t][yy] += 1
fig, ax = plt.subplots(figsize=(10, 5))
years_all = sorted({y for c in topic_year.values() for y in c})
bottom = np.zeros(len(years_all))
for t in range(n_topics):
    vals = np.array([topic_year[t].get(y, 0) for y in years_all])
    ax.bar(years_all, vals, bottom=bottom, label=f"T{t+1}: {topic_terms[t][0]}",
           color=plt.get_cmap("tab10")(t))
    bottom += vals
ax.legend(fontsize=8); ax.set_xlabel("Year"); ax.set_ylabel("Papers")
ax.set_title("2.2b Topic distribution over time")
plt.tight_layout(); plt.savefig(P2 / "2.2b_topic_year_trend.png", dpi=130); plt.close()
print(f"  2.2b → topic_year_trend.png")

# 2.3 Geographic — country from affiliations
countries = Counter()
for r in rows:
    aff = col(r, "affiliations") or ""
    if not aff: continue
    for unit in str(aff).split("|"):
        bits = [b.strip() for b in unit.split(";")]
        if len(bits) >= 3:
            c = bits[-1]
            c = c.replace("People's Republic of China", "China")
            if c: countries[c] += 1
top_c = countries.most_common(15)
fig, ax = plt.subplots(figsize=(9, 6))
ax.barh([c for c, _ in top_c[::-1]], [n for _, n in top_c[::-1]], color="#4CAF50")
ax.set_xlabel("Author affiliations"); ax.set_title(f"2.3 Top countries (n={sum(countries.values())} aff slots)")
plt.tight_layout(); plt.savefig(P2 / "2.3_geographic.png", dpi=130); plt.close()
print(f"  2.3 → geographic.png  ({len(countries)} countries)")

# ============================================================
# Phase 3 — Visuals
# ============================================================

try:
    from wordcloud import WordCloud
    stop = {"china","chinese","steel","iron","carbon","emission","emissions","study","based","results","industry","using","paper","effect"}
    text_blob = " ".join(
        f"{col(r,'title') or ''} {col(r,'abstract') or ''} {col(r,'author_keywords') or ''}"
        for r in rows
    )
    wc = WordCloud(width=1200, height=600, background_color="white",
                   stopwords=stop, max_words=120, colormap="viridis").generate(text_blob)
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.imshow(wc); ax.axis("off"); ax.set_title("3.1 Wordcloud — titles + abstracts + keywords")
    plt.tight_layout(); plt.savefig(P3 / "3.1_wordcloud_overall.png", dpi=130); plt.close()
    print(f"  3.1 → wordcloud_overall.png")

    if has_topic:
        topics_set = sorted({str(col(r, "topic_label") or "unlabelled") for r in rows})
        for ti, t in enumerate(topics_set, start=1):
            text_t = " ".join(
                f"{col(r,'title') or ''} {col(r,'abstract') or ''} {col(r,'author_keywords') or ''}"
                for r in rows if str(col(r, "topic_label") or "unlabelled") == t
            )
            if not text_t.strip(): continue
            wct = WordCloud(width=900, height=500, background_color="white",
                            stopwords=stop, max_words=80, colormap="plasma").generate(text_t)
            fig, ax = plt.subplots(figsize=(9, 5))
            ax.imshow(wct); ax.axis("off"); ax.set_title(f"3.2.{ti} Wordcloud — {t}")
            plt.tight_layout(); plt.savefig(P3 / f"3.2.{ti}_wordcloud_{t}.png", dpi=130); plt.close()
            print(f"  3.2.{ti} → wordcloud_{t}.png")
except ImportError:
    print("  ! wordcloud not installed, skipping phase 3")

print(f"\n[done] charts in:\n  {P1}/\n  {P2}/\n  {P3}/")
