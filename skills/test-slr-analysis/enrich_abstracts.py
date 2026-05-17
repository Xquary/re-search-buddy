"""Post-hoc abstract enrichment for existing SLR xlsx.

Calls the Scopus Abstract Retrieval API for every paper missing an abstract,
using the scopus_id already present in the xlsx.  Operates across all sheets
that carry both a `scopus_id` and an `abstract` column.

Usage:
    PYTHONPATH=src .venv/bin/python skills/test-slr-analysis/enrich_abstracts.py
"""

from __future__ import annotations

import asyncio
import os
import sys
from pathlib import Path

import httpx
import openpyxl
from dotenv import load_dotenv

load_dotenv()

XLSX_PATH = "output/Seeds of Green_Methodology/SLR_Scopus_ABM_Energy_2016-2026_qcomb_max2000/SLR_Scopus_ABM_Energy_2016-2026_qcomb_max2000.xlsx"
ABSTRACT_URL = "https://api.elsevier.com/content/abstract/scopus_id/{}"

SCOPUS_API_KEY = os.environ.get("SCOPUS_API_KEY", "")
if not SCOPUS_API_KEY:
    print("SCOPUS_API_KEY not set — aborting.")
    sys.exit(1)

HEADERS = {
    "X-ELS-APIKey": SCOPUS_API_KEY,
    "Accept": "application/json",
}


def discover_sheets(wb: openpyxl.Workbook) -> list[tuple[str, int, int]]:
    """Return (sheet_name, scopus_id_col_1based, abstract_col_1based) for relevant sheets."""
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        if "scopus_id" in headers and "abstract" in headers:
            sid_col = headers.index("scopus_id") + 1
            abs_col = headers.index("abstract") + 1
            sheets.append((name, sid_col, abs_col))
    return sheets


def find_missing(
    wb: openpyxl.Workbook, sheets_info: list[tuple[str, int, int]]
) -> dict[str, list[tuple[str, int]]]:
    """Collect unique (scopus_id) → list of (sheet_name, row) for papers with empty abstract."""
    scopus_rows: dict[str, list[tuple[str, int]]] = {}
    for sheet_name, sid_col, abs_col in sheets_info:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            sid = ws.cell(row, sid_col).value
            abstract = ws.cell(row, abs_col).value
            if sid and not abstract:
                sid = str(sid).strip()
                scopus_rows.setdefault(sid, []).append((sheet_name, row))
    return scopus_rows


async def fetch_abstracts(
    scopus_ids: set[str],
) -> dict[str, str | None]:
    """Fetch abstracts for a deduplicated set of scopus_ids.  Returns {id: abstract | None}."""
    results: dict[str, str | None] = {}
    sem = asyncio.Semaphore(2)  # concurrency cap to stay under 9 req/s

    async def _fetch_one(client: httpx.AsyncClient, sid: str):
        async with sem:
            await asyncio.sleep(0.15)  # inter-request delay
            try:
                resp = await client.get(
                    ABSTRACT_URL.format(sid),
                    params={"view": "FULL"},
                )
                resp.raise_for_status()
                data = resp.json()
                root = (
                    data.get("abstracts-retrieval-response")
                    or data.get("abstract-retrieval-response")
                    or {}
                )
                abstract = root.get("coredata", {}).get("dc:description")
                results[sid] = abstract
            except Exception as exc:
                print(f"  [warn] {sid}: {exc}")
                results[sid] = None

    async with httpx.AsyncClient(
        headers=HEADERS, timeout=30.0, follow_redirects=True
    ) as client:
        tasks = [_fetch_one(client, sid) for sid in scopus_ids]
        await asyncio.gather(*tasks)

    return results


def apply_abstracts(
    wb: openpyxl.Workbook,
    scopus_rows: dict[str, list[tuple[str, int]]],
    abstracts: dict[str, str | None],
    sheets_info: list[tuple[str, int, int]],
) -> int:
    """Write fetched abstracts back into the xlsx. Returns number of cells updated."""
    updated = 0
    # Build quick lookup of abstract column per sheet
    abs_col_map = {name: abs_col for name, _, abs_col in sheets_info}
    for sid, rows in scopus_rows.items():
        abstract = abstracts.get(sid)
        if abstract:
            for sheet_name, row in rows:
                ws = wb[sheet_name]
                ws.cell(row, abs_col_map[sheet_name], value=abstract)
                updated += 1
    return updated


async def main():
    if not os.path.exists(XLSX_PATH):
        print(f"xlsx not found: {XLSX_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH)
    sheets_info = discover_sheets(wb)
    print(f"Sheets with scopus_id+abstract: {[s[0] for s in sheets_info]}")

    scopus_rows = find_missing(wb, sheets_info)
    unique_ids = set(scopus_rows.keys())
    total_missing = sum(len(v) for v in scopus_rows.values())
    print(f"Papers missing abstracts: {total_missing} (unique scopus_ids: {len(unique_ids)})")

    if not unique_ids:
        print("Nothing to do — all papers already have abstracts.")
        wb.close()
        return

    print(f"Fetching {len(unique_ids)} abstracts via Scopus Abstract Retrieval API…")
    abstracts = await fetch_abstracts(unique_ids)

    hit = sum(1 for v in abstracts.values() if v)
    miss = len(unique_ids) - hit
    print(f"Fetched: {hit} success, {miss} failures")

    updated = apply_abstracts(wb, scopus_rows, abstracts, sheets_info)
    print(f"Cells updated: {updated}")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"Saved: {XLSX_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
