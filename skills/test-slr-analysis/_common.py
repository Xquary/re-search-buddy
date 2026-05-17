"""Shared loader for all analysis_*.py scripts.

CLI contract:
  --xlsx <path>                 required
  --sheet <name>                optional; default = active
  --topic-col <col>             optional; default 'keyword_hit' (legacy) or 'topic_label'
  --exclude-journals "A|B|C"    optional; pipe-separated, case-insensitive
  --phase-dir <path>            optional; override charts/analysis root
  --subdir <name>               optional; sub-folder under phase-dir (e.g. 'keyword_clean')

Returns a context dict consumed by every analysis script. Auto-builds subset_map /
order / labels / colors from the unique non-empty values in the topic column,
so the legacy 'D-energy/E-industrial' style works AND our 'SOE/SupplyChain/Policy/PoliticalEcon' works.
"""
import argparse
from pathlib import Path
import openpyxl

PALETTE = ["#2196F3", "#FF9800", "#4CAF50", "#9C27B0",
           "#E91E63", "#00BCD4", "#FFC107", "#795548"]


def parse_and_load():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--topic-col", default="keyword_hit")
    ap.add_argument("--exclude-journals", default="")
    ap.add_argument("--phase-dir", default=None)
    ap.add_argument("--subdir", default="")
    args, _ = ap.parse_known_args()  # tolerate extras (e.g. --top-cited for citations)

    xlsx = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = [[ws.cell(row=r, column=i+1).value for i in range(len(headers))]
            for r in range(2, ws.max_row + 1)]

    raw_n = len(rows)
    excl = {j.strip().lower() for j in args.exclude_journals.split("|") if j.strip()}
    if excl and "journal" in idx:
        ji = idx["journal"]
        rows = [r for r in rows if str(r[ji] or "").strip().lower() not in excl]
        print(f"[filter] dropped {raw_n - len(rows)} papers; {len(rows)} remain")

    topic_col = args.topic_col if args.topic_col in idx else (
        "topic_label" if "topic_label" in idx else None
    )

    topics: list[str] = []
    if topic_col:
        ti = idx[topic_col]
        seen = []
        for r in rows:
            v = r[ti]
            if not v: continue
            # support multi-value cells like "D-energy; E-industrial"
            for piece in str(v).replace(",", ";").split(";"):
                p = piece.strip()
                if p and p not in seen:
                    seen.append(p)
        topics = seen

    subset_order = topics
    subset_map = {t: t for t in topics}
    subset_labels = {t: t for t in topics}
    subset_colors = {t: PALETTE[i % len(PALETTE)] for i, t in enumerate(topics)}

    base_dir = Path(args.phase_dir) if args.phase_dir else xlsx.parent / "charts" / "analysis"
    if args.subdir:
        base_dir = base_dir / args.subdir

    return {
        "args": args,
        "xlsx_path": xlsx,
        "sheet_name": ws.title,
        "headers": headers,
        "idx": idx,
        "rows": rows,
        "topics": topics,
        "topic_col": topic_col,
        "subset_map": subset_map,
        "subset_order": subset_order,
        "subset_labels": subset_labels,
        "subset_colors": subset_colors,
        "base_dir": base_dir,
    }


def get(row, idx, name, default=None):
    return row[idx[name]] if name in idx and row[idx[name]] is not None else default
