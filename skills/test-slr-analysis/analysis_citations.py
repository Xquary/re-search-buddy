"""Phase 1.2 — Citation analysis: distribution + boxplot + top-cited + score-vs-cites with highlights.

Highlights = union(top-N most cited, top-M most relevant). Labels via adjustText.
Also writes a *_highlighted helper sheet to the xlsx with a 'selected_reason' column.
"""
import argparse, sys
from pathlib import Path
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import numpy as np
import openpyxl
from _common import parse_and_load, get

# Extra args (parsed via known/unknown to coexist with _common's argparse-consuming load)
# _common.parse_and_load runs argparse; we re-parse for the extras after.
_extra = argparse.ArgumentParser(add_help=False)
_extra.add_argument("--top-cited", type=int, default=15)
_extra.add_argument("--top-relevant", type=int, default=15)
_extra.add_argument("--write-highlight-sheet", action="store_true", default=True)
_extra.add_argument("--no-write-highlight-sheet", dest="write_highlight_sheet", action="store_false")
_known, _ = _extra.parse_known_args()
TOP_CITED = _known.top_cited
TOP_REL = _known.top_relevant
WRITE_SHEET = _known.write_highlight_sheet

ctx = parse_and_load()
rows, idx, headers = ctx["rows"], ctx["idx"], ctx["headers"]
subset_order = ctx["subset_order"]
subset_labels = ctx["subset_labels"]
subset_colors = ctx["subset_colors"]
has_kw = "keyword_hit" in idx and bool(subset_order)
out_dir = ctx["base_dir"] / "1_profiling"; out_dir.mkdir(parents=True, exist_ok=True)

papers = []
for ri, r in enumerate(rows):
    cites_raw = get(r, idx, "citations", "")
    try: cites = int(str(cites_raw).strip()) if str(cites_raw).strip().lstrip("-").isdigit() else 0
    except: cites = 0
    try: score = float(get(r, idx, "score", 0) or 0)
    except: score = 0.0
    fa = str(get(r, idx, "first_author", "") or "")
    yr = str(get(r, idx, "year", "") or "")[:4]
    dom = "none"
    if has_kw:
        hits = str(get(r, idx, "keyword_hit", "") or "")
        for s in subset_order:
            if s in hits: dom = s; break
    papers.append({
        "row_idx": ri, "citations": cites, "score": score,
        "first_author": fa, "year": yr, "dominant": dom,
        "title": str(get(r, idx, "title", "") or ""),
    })

all_cites = [p["citations"] for p in papers] or [0]
print(f"[{ctx['sheet_name']}] n={len(papers)} cites mean={np.mean(all_cites):.1f} median={np.median(all_cites):.0f} max={max(all_cites)}")

# pick highlights = union(top cited, top relevant)
by_cite = sorted(papers, key=lambda p: p["citations"], reverse=True)[:TOP_CITED]
by_rel = sorted(papers, key=lambda p: p["score"], reverse=True)[:TOP_REL]
cited_ids = {id(p) for p in by_cite}
rel_ids = {id(p) for p in by_rel}
highlighted = []
seen = set()
for p in by_cite + by_rel:
    if id(p) in seen: continue
    seen.add(id(p))
    reasons = []
    if id(p) in cited_ids: reasons.append("high cited")
    if id(p) in rel_ids: reasons.append("high relevant")
    p["selected_reason"] = ", ".join(reasons)
    highlighted.append(p)
print(f"  highlights: {len(highlighted)} (top {TOP_CITED} cited ∪ top {TOP_REL} relevant)")

def ay(p):
    last = p["first_author"].split(",")[0].strip() if p["first_author"] else "?"
    return f"{last} ({p['year'] or '?'})"

if has_kw:
    fig, axes_2d = plt.subplots(2, 2, figsize=(16, 13))
    axes = [axes_2d[0][0], axes_2d[0][1], axes_2d[1][0], axes_2d[1][1]]
else:
    fig, axes_1d = plt.subplots(1, 2, figsize=(16, 6))
    axes = [axes_1d[0], axes_1d[1], None, None]

# 1.2a Distribution
ax = axes[0]
bins = np.logspace(0, np.log10(max(all_cites) + 1), 25)
ax.hist([max(c, 0.5) for c in all_cites], bins=bins, color="#546E7A", edgecolor="white", alpha=0.85)
ax.set_xscale("log"); ax.set_xlabel("Citations (log)", fontsize=13); ax.set_ylabel("Papers", fontsize=13)
ax.set_title(f"1.2a Citation distribution (mean={np.mean(all_cites):.1f}, median={np.median(all_cites):.0f})",
             fontsize=14, fontweight="bold")

# 1.2b Boxplot by subset
if has_kw:
    ax = axes[1]
    sd = {s: [p["citations"] for p in papers if p["dominant"] == s] for s in subset_order}
    sd = {s: v for s, v in sd.items() if v}
    if sd:
        bp = ax.boxplot(list(sd.values()), labels=list(sd.keys()), patch_artist=True)
        for patch, s in zip(bp["boxes"], sd.keys()):
            patch.set_facecolor(subset_colors[s]); patch.set_alpha(0.6)
    ax.set_ylabel("Citations", fontsize=13)
    ax.set_title("1.2b Citations by subset", fontsize=14, fontweight="bold")

# 1.2c Top-15 most cited (always)
ax = axes[2] if has_kw else axes[1]
top15 = sorted(papers, key=lambda p: p["citations"], reverse=True)[:15]
labels = [ay(p) for p in top15]
cs = [p["citations"] for p in top15]
colors = [subset_colors.get(p["dominant"], "#546E7A") for p in top15] if has_kw else "#546E7A"
bars = ax.barh(range(len(labels)), cs, color=colors, edgecolor="white", height=0.7)
ax.set_yticks(range(len(labels))); ax.set_yticklabels(labels, fontsize=11)
ax.invert_yaxis(); ax.set_xlabel("Citations", fontsize=13)
ax.set_title("1.2c Top-15 most cited", fontsize=14, fontweight="bold")
for bar, v in zip(bars, cs):
    ax.text(bar.get_width()+1, bar.get_y()+bar.get_height()/2, str(v),
            va="center", fontsize=10, fontweight="bold")

# 1.2d Score vs citations with highlights + adjustText labels
if has_kw:
    ax = axes[3]
    hi_set = set(id(p) for p in highlighted)
    # Background (non-highlighted)
    for s in subset_order:
        sub = [p for p in papers if p["dominant"] == s and id(p) not in hi_set]
        if sub:
            ax.scatter([p["score"] for p in sub], [p["citations"] for p in sub],
                       c=subset_colors[s], alpha=0.3, s=25, edgecolors="none")
    # Foreground (highlighted)
    for p in highlighted:
        c = subset_colors.get(p["dominant"], "#9E9E9E")
        ax.scatter(p["score"], p["citations"], c=c, alpha=0.95, s=140,
                   edgecolors="black", linewidth=1.4, zorder=5)

    # Labels via adjustText (graceful fallback if missing)
    texts = []
    for p in highlighted:
        texts.append(ax.text(p["score"], p["citations"], ay(p),
                             fontsize=9, fontweight="bold", alpha=0.9,
                             bbox=dict(boxstyle="round,pad=0.15", facecolor="white",
                                       alpha=0.85, edgecolor="#CFD8DC")))
    try:
        from adjustText import adjust_text
        adjust_text(texts, ax=ax,
                    arrowprops=dict(arrowstyle="->", color="#78909C", lw=0.5),
                    expand_text=(1.2, 1.4), expand_points=(1.2, 1.4))
    except Exception as e:
        print(f"  [warn] adjustText failed ({e}); using static labels")

    ax.set_xlabel("Cosine similarity score", fontsize=13)
    ax.set_ylabel("Citations", fontsize=13)
    ax.set_title(f"1.2d Score vs. citations  (● top-{TOP_CITED} cited ∪ top-{TOP_REL} relevant = {len(highlighted)} labeled)",
                 fontsize=13, fontweight="bold")
    ax.legend(handles=[Patch(facecolor=subset_colors[s], label=subset_labels[s]) for s in subset_order],
              fontsize=10, loc="lower right")

plt.suptitle(f"{ctx['sheet_name']} — Citation Analysis", fontsize=17, fontweight="bold", y=1.02)
plt.tight_layout()
fig.savefig(out_dir / "1.2_citations.png", dpi=150, bbox_inches="tight")
plt.close()
print(f"  saved {out_dir / '1.2_citations.png'}")

# Write *_highlighted helper sheet to xlsx
if WRITE_SHEET:
    xlsx_path = ctx["xlsx_path"]
    wb = openpyxl.load_workbook(xlsx_path)
    src_sheet = ctx["sheet_name"]
    short = {"keyword_verified_clean": "kw_clean", "missing_topic_only_clean": "missing"}.get(src_sheet, src_sheet[:20])
    out_name = f"{short}_highlighted"
    if out_name in wb.sheetnames: del wb[out_name]
    out_ws = wb.create_sheet(out_name)
    out_headers = list(headers) + (["selected_reason"] if "selected_reason" not in headers else [])
    out_ws.append(out_headers)
    hi_sorted = sorted(highlighted, key=lambda p: p["citations"], reverse=True)
    for p in hi_sorted:
        row = list(rows[p["row_idx"]])
        # ensure row is the right length (pad if loader added keyword_hit alias)
        while len(row) < len(headers): row.append("")
        out_ws.append(row + [p["selected_reason"]])
    wb.save(xlsx_path)
    print(f"  wrote sheet '{out_name}' ({len(highlighted)} rows) to xlsx")
