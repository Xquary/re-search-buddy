"""Adds two cleaned sheets to an SLR xlsx based on topic keyword buckets:

  - keyword_verified_clean   : papers whose title/abstract/author_keywords contain
                                at least one term from any bucket. Adds 'keyword_hit'
                                column listing matched buckets (semicolon-separated).
  - missing_topic_only_clean : papers retrieved by Scopus but whose text did NOT
                                match any bucket — likely retrieved on a tangential term.

Idempotent: re-running overwrites both sheets in-place.

Usage:
  PYTHONPATH=src .venv/bin/python skills/test-slr-analysis/enrich_keyword_clean.py \\
    --xlsx <path> --buckets-yaml <path>   # OR pass --buckets-json
"""
import argparse, json, sys
from pathlib import Path
import openpyxl

ap = argparse.ArgumentParser()
ap.add_argument("--xlsx", required=True)
ap.add_argument("--buckets-json", default=None,
                help='inline JSON: {"SOE":["state-owned","SOE",...], ...}')
ap.add_argument("--buckets-yaml", default=None,
                help="path to YAML file with same structure")
ap.add_argument("--source-sheet", default=None,
                help="source sheet name; default = first sheet")
args = ap.parse_args()

if args.buckets_yaml:
    import yaml
    with open(args.buckets_yaml) as f:
        BUCKETS = yaml.safe_load(f)
elif args.buckets_json:
    BUCKETS = json.loads(args.buckets_json)
else:
    print("error: pass --buckets-json or --buckets-yaml")
    sys.exit(1)

BUCKETS_LC = {k: [t.lower() for t in v] for k, v in BUCKETS.items()}

wb = openpyxl.load_workbook(args.xlsx)
src = wb[args.source_sheet] if args.source_sheet else wb.worksheets[0]
headers = [c.value for c in src[1]]
idx = {h: i for i, h in enumerate(headers)}
need = {"title", "abstract"}
missing = need - set(idx)
if missing:
    print(f"error: source sheet missing columns {missing}")
    sys.exit(1)

ti, ai = idx["title"], idx["abstract"]
ki = idx.get("author_keywords")

# remove existing output sheets so we re-create them clean
for name in ["keyword_verified_clean", "missing_topic_only_clean"]:
    if name in wb.sheetnames:
        del wb[name]

ver = wb.create_sheet("keyword_verified_clean")
mis = wb.create_sheet("missing_topic_only_clean")
ver_headers = headers + (["keyword_hit"] if "keyword_hit" not in idx else [])
ver.append(ver_headers)
mis.append(headers)

n_ver = 0
n_mis = 0
bucket_counts = {b: 0 for b in BUCKETS_LC}
for r in range(2, src.max_row + 1):
    row = [src.cell(row=r, column=i+1).value for i in range(len(headers))]
    blob_parts = [str(row[ti] or ""), str(row[ai] or "")]
    if ki is not None: blob_parts.append(str(row[ki] or ""))
    blob = " ".join(blob_parts).lower()

    hits = []
    for bucket, terms in BUCKETS_LC.items():
        if any(t in blob for t in terms):
            hits.append(bucket)
            bucket_counts[bucket] += 1

    if hits:
        out_row = row + ["; ".join(hits)] if "keyword_hit" not in idx else list(row)
        # if keyword_hit existed, overwrite it
        if "keyword_hit" in idx:
            out_row[idx["keyword_hit"]] = "; ".join(hits)
        ver.append(out_row)
        n_ver += 1
    else:
        mis.append(row)
        n_mis += 1

wb.save(args.xlsx)
print(f"[verified] {n_ver} papers  → sheet 'keyword_verified_clean'")
for b, c in bucket_counts.items():
    print(f"     {b:20s} {c}")
print(f"[missing ] {n_mis} papers  → sheet 'missing_topic_only_clean'")
print(f"[done] {Path(args.xlsx).name}")
